# src/export/excel_exporter.py
# Exports generated test cases to Excel in your organization's exact format.
# Produces .xlsx with proper columns, styling, and sheet structure.

from pathlib import Path
from datetime import datetime
from typing import List, Dict
from loguru import logger

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

import config


# ── Your org's exact column headers ─────────────────────────
COLUMNS = [
    "Folder",
    "Name",
    "Precondition",
    "Steps",
    "Test Data",
    "Expected Result",
    "Actual Result",
    "Priority",
    "Status"
]

# Column widths (in characters)
COLUMN_WIDTHS = [35, 55, 45, 55, 25, 45, 20, 10, 12]

# Header styling — blue background, white bold text
HEADER_FILL  = PatternFill("solid", start_color="1F4E79")
HEADER_FONT  = Font(bold=True, color="FFFFFF", name="Arial", size=10)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Row styling
ROW_FONT     = Font(name="Arial", size=9)
ROW_ALIGN    = Alignment(vertical="top", wrap_text=True)

# Priority color coding
PRIORITY_COLORS = {
    "High":   "FFE0E0",   # light red
    "Medium": "FFF3CD",   # light yellow
    "Low":    "E8F5E9",   # light green
}

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)


def parse_llm_test_cases(raw_text: str) -> List[Dict]:
    """
    Parses LLM output (pipe-separated format) into list of dicts.
    Each dict maps to one row in the Excel sheet.
    Skips malformed lines gracefully.
    """
    rows = []
    for line in raw_text.strip().split("\n"):
        line = line.strip()
        # Skip empty lines, headers, separators
        if not line or line.startswith("#") or line.startswith("=") or "|" not in line:
            continue
        parts = [p.strip() for p in line.split("|")]
        if len(parts) < 6:
            continue
        rows.append({
            "Folder":          parts[0] if len(parts) > 0 else "",
            "Name":            parts[1] if len(parts) > 1 else "",
            "Precondition":    parts[2] if len(parts) > 2 else "",
            "Steps":           parts[3] if len(parts) > 3 else "",
            "Test Data":       parts[4] if len(parts) > 4 else "N/A",
            "Expected Result": parts[5] if len(parts) > 5 else "",
            "Actual Result":   "",        # always blank — filled during execution
            "Priority":        parts[6] if len(parts) > 6 else "Medium",
            "Status":          ""         # blank by default
        })
    return rows


def style_header_row(sheet, row_num: int):
    """Apply header styling to a row."""
    for col_num in range(1, len(COLUMNS) + 1):
        cell = sheet.cell(row=row_num, column=col_num)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border    = THIN_BORDER


def style_data_row(sheet, row_num: int, priority: str):
    """Apply data row styling with priority color."""
    fill_color = PRIORITY_COLORS.get(priority, "FFFFFF")
    row_fill   = PatternFill("solid", start_color=fill_color)

    for col_num in range(1, len(COLUMNS) + 1):
        cell = sheet.cell(row=row_num, column=col_num)
        cell.font      = ROW_FONT
        cell.alignment = ROW_ALIGN
        cell.fill      = row_fill
        cell.border    = THIN_BORDER


def create_sheet(wb, sheet_name: str, rows: List[Dict]):
    """
    Creates one sheet in the workbook with the given test case rows.
    sheet_name: e.g. "DABI" or "PROGOTI" or your project name
    """
    ws = wb.create_sheet(title=sheet_name)

    # ── Write header row ─────────────────────────────────────
    for col_num, col_name in enumerate(COLUMNS, start=1):
        ws.cell(row=1, column=col_num, value=col_name)
    style_header_row(ws, 1)
    ws.row_dimensions[1].height = 25

    # ── Write data rows ──────────────────────────────────────
    for row_num, row_data in enumerate(rows, start=2):
        for col_num, col_name in enumerate(COLUMNS, start=1):
            ws.cell(row=row_num, column=col_num, value=row_data.get(col_name, ""))
        priority = row_data.get("Priority", "Medium")
        style_data_row(ws, row_num, priority)
        ws.row_dimensions[row_num].height = 55

    # ── Set column widths ────────────────────────────────────
    for col_num, width in enumerate(COLUMN_WIDTHS, start=1):
        col_letter = openpyxl.utils.get_column_letter(col_num)
        ws.column_dimensions[col_letter].width = width

    # ── Freeze header row ────────────────────────────────────
    ws.freeze_panes = "A2"

    logger.info(f"Sheet '{sheet_name}' created with {len(rows)} test cases.")
    return ws


def export_to_excel(
    raw_llm_output: str,
    cr_id: str = "",
    project_name: str = "Project",
    filename: str = None
) -> Path:
    """
    Main export function.
    Takes raw LLM output and saves as .xlsx in your org's format.

    raw_llm_output: the pipe-separated test cases from LLM
    cr_id:          e.g. "CR-2024-089" — used in filename
    project_name:   used as sheet name (e.g. "DABI", "PROGOTI")
    filename:       optional custom filename
    """
    # Parse LLM output into rows
    rows = parse_llm_test_cases(raw_llm_output)
    if not rows:
        logger.warning("No valid test cases parsed from LLM output.")
        rows = []

    # Create workbook
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Create project sheet
    create_sheet(wb, project_name.upper(), rows)

    # Build filename
    if not filename:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        cr_part   = f"{cr_id}_" if cr_id else ""
        filename  = f"test_cases_{cr_part}{timestamp}.xlsx"

    output_path = config.EXPORTS_PATH / filename
    wb.save(str(output_path))
    logger.info(f"Excel exported to: {output_path} ({len(rows)} test cases)")
    return output_path


# ── Quick test ───────────────────────────────────────────────
if __name__ == "__main__":
    sample_output = """
/OTC/Loan Management/Repayment | Verify weekly option in repayment frequency dropdown | BAO user is logged in; Loan module configured; Valid loan account exists | 1. Open Loan Management. 2. Navigate to Repayment. 3. Open Frequency dropdown. | N/A | Weekly option is visible and selectable. | High
/OTC/Loan Management/Repayment | Verify installment amount recalculates for weekly frequency | BAO user logged in; Loan with monthly schedule exists | 1. Open existing loan. 2. Change frequency to weekly. 3. Save. 4. Check installment amount. | Loan ID: L-001; Monthly amount: 5000 | Installment amount recalculates correctly for weekly schedule. | High
/OTC/Loan Management/Repayment | Verify weekly repayment schedule cannot be saved without due date | BAO user logged in; Loan module configured | 1. Open Loan. 2. Select weekly frequency. 3. Leave due date empty. 4. Save. | N/A | System shows validation error: Due date is required. | High
/OTC/Notifications/SMS | Verify SMS sent when weekly payment is due | SMS gateway configured; Loan with weekly schedule exists | 1. Set up weekly loan. 2. Trigger payment due event. 3. Check SMS logs. | Borrower phone: 01XXXXXXXXX | SMS notification is sent to borrower on weekly payment due date. | Medium
    """

    path = export_to_excel(
        raw_llm_output=sample_output,
        cr_id="CR-2024-089",
        project_name="DABI"
    )
    print(f"\nExcel file created: {path}")
    print("Open it in LibreOffice or Excel to verify the format!")